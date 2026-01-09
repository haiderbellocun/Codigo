# -*- coding: utf-8 -*-
"""
AWS Lambda: disparador de AWS Transcribe por evento S3.

Flujo:
- Se dispara cuando llega un audio a S3 (ObjectCreated).
- Extrae un correo desde el S3 key (regex).
- Valida si el correo está aprobado contra un Excel alojado en S3.
- Si está aprobado, inicia un job de AWS Transcribe con salida en OUT_BUCKET/OUT_PREFIX/YYYY-MM-DD/Grabaciones/.

Notas:
- Este handler NO descarga el audio; Transcribe lo lee directamente desde S3.
- Debes permitir que el servicio `transcribe.amazonaws.com` lea el bucket de entrada
  y escriba en el bucket de salida (ver docs/IAM.md).
"""

from __future__ import annotations

import os
import re
import json
import urllib.parse
from datetime import datetime
from io import BytesIO
from typing import Optional, Set, Dict, Any, List

import boto3
from botocore.exceptions import ClientError

try:
    # Más liviano que pandas/numpy para Lambda ZIP
    from openpyxl import load_workbook
    _HAS_OPENPYXL = True
except Exception:
    _HAS_OPENPYXL = False


# ===================== CONFIG (variables de entorno) =====================
REGION = os.environ.get("REGION", "us-east-1")

# Dónde escribir el resultado (Transcribe output)
OUT_BUCKET = os.environ.get("OUT_BUCKET", "")
OUT_PREFIX = os.environ.get("OUT_PREFIX", "txt/prue")

# Config transcribe
LANGUAGE = os.environ.get("LANGUAGE", "auto")  # "auto" o ej. "es-US"
SPEAKERS = int(os.environ.get("SPEAKERS", "0") or "0")  # >=2 activa speaker labels

# Excel con correos aprobados
BASE_VENTAS_BUCKET = os.environ.get("BASE_VENTAS_BUCKET", "")
BASE_VENTAS_KEY = os.environ.get("BASE_VENTAS_KEY", "")
APPROVED_SHEET = os.environ.get("APPROVED_SHEET", "")  # opcional: nombre de hoja
EMAIL_COLUMNS = os.environ.get("EMAIL_COLUMNS", "correo,email").split(",")

# Control de extensiones (opcional)
ALLOWED_EXTS = [e.strip().lower() for e in os.environ.get("ALLOWED_EXTS", "wav,mp3,m4a,flac,wma").split(",") if e.strip()]

# Regex de correo (puedes endurecerlo si lo necesitas)
EMAIL_REGEX = os.environ.get("EMAIL_REGEX", r'[\w.+-]+@[\w-]+\.[\w.-]+')

# ===================== AWS CLIENTS =====================
transcribe = boto3.client("transcribe", region_name=REGION)
s3_client = boto3.client("s3", region_name=REGION)

# ===================== CACHE (entre invocaciones dentro del mismo container) =====================
APROBADOS_CACHE: Optional[Set[str]] = None
APROBADOS_ETAG: Optional[str] = None


# ===================== Helpers =====================
def extract_email_from_key(key: str) -> Optional[str]:
    match = re.search(EMAIL_REGEX, key)
    return match.group(0).lower() if match else None


def sanitize_job_name(key: str) -> str:
    # Transcribe JobName: alfanumérico + . _ - (seamos defensivos)
    base = re.sub(r"[^A-Za-z0-9._-]+", "-", os.path.basename(key))
    ts = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
    return f"tr-{ts}-{base}"[:200]


def _get_excel_etag(bucket: str, key: str) -> Optional[str]:
    try:
        h = s3_client.head_object(Bucket=bucket, Key=key)
        return h.get("ETag")
    except Exception:
        return None


def get_lista_aprobados() -> Set[str]:
    """
    Carga correos aprobados desde Excel en S3.
    Cachea el resultado entre invocaciones (warm starts) y refresca si cambia el ETag.
    """
    global APROBADOS_CACHE, APROBADOS_ETAG

    if not BASE_VENTAS_BUCKET or not BASE_VENTAS_KEY:
        print("ERROR: BASE_VENTAS_BUCKET/BASE_VENTAS_KEY no configuradas.")
        return set()

    if not _HAS_OPENPYXL:
        print("ERROR: openpyxl no está instalado. Instálalo (requirements.txt) o usa container.")
        return set()

    etag = _get_excel_etag(BASE_VENTAS_BUCKET, BASE_VENTAS_KEY)
    if APROBADOS_CACHE is not None and (etag is None or etag == APROBADOS_ETAG):
        return APROBADOS_CACHE

    print("COLD/WARM REFRESH: Cargando lista de correos aprobados desde Excel en S3...")

    try:
        obj = s3_client.get_object(Bucket=BASE_VENTAS_BUCKET, Key=BASE_VENTAS_KEY)
        data = obj["Body"].read()

        wb = load_workbook(filename=BytesIO(data), read_only=True, data_only=True)
        if APPROVED_SHEET and APPROVED_SHEET in wb.sheetnames:
            ws = wb[APPROVED_SHEET]
        else:
            ws = wb[wb.sheetnames[0]]

        # Leer cabecera (primera fila)
        rows = ws.iter_rows(min_row=1, max_row=1, values_only=True)
        header = next(rows, None)
        if not header:
            raise ValueError("Excel sin cabecera (fila 1 vacía).")

        header_norm = [str(c).strip().lower() if c is not None else "" for c in header]

        # Buscar columna de correo
        email_col_idx = None
        for candidate in [c.strip().lower() for c in EMAIL_COLUMNS if c.strip()]:
            if candidate in header_norm:
                email_col_idx = header_norm.index(candidate)
                break
        if email_col_idx is None:
            raise ValueError(f"No se encontró columna de correo. Esperadas: {EMAIL_COLUMNS}")

        correos: Set[str] = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            val = row[email_col_idx] if email_col_idx < len(row) else None
            if val is None:
                continue
            s = str(val).strip().lower()
            if s:
                correos.add(s)

        print(f"Lista cargada: {len(correos)} correos aprobados.")
        APROBADOS_CACHE = correos
        APROBADOS_ETAG = etag
        return correos

    except Exception as e:
        print(f"ERROR CRÍTICO: No se pudo cargar lista de aprobación. Error: {e}")
        APROBADOS_CACHE = set()
        APROBADOS_ETAG = etag
        return APROBADOS_CACHE


def _build_output_path() -> str:
    nombre_carpeta_diaria = datetime.utcnow().strftime("%Y-%m-%d")
    # OUT_PREFIX/YYYY-MM-DD/Grabaciones
    return os.path.join(OUT_PREFIX, nombre_carpeta_diaria, "Grabaciones")


def _is_allowed_key(key: str) -> bool:
    ext = key.split(".")[-1].lower() if "." in key else ""
    return ext in ALLOWED_EXTS


# ===================== HANDLER =====================
def lambda_handler(event: Dict[str, Any], context: Any) -> Dict[str, Any]:
    aprobados = get_lista_aprobados()
    if not aprobados:
        return {"statusCode": 200, "body": "Lista de aprobación vacía o no cargada."}

    records = event.get("Records") or []
    if not records:
        return {"statusCode": 200, "body": "Evento sin Records."}

    processed: List[Dict[str, Any]] = []
    skipped: List[Dict[str, Any]] = []

    for record in records:
        try:
            in_bucket = record["s3"]["bucket"]["name"]
            in_key = urllib.parse.unquote_plus(record["s3"]["object"]["key"], encoding="utf-8")

            if not _is_allowed_key(in_key):
                skipped.append({"key": in_key, "reason": "ext_not_allowed"})
                continue

            correo = extract_email_from_key(in_key)
            if not correo:
                skipped.append({"key": in_key, "reason": "no_email_in_key"})
                continue

            if correo not in aprobados:
                skipped.append({"key": in_key, "correo": correo, "reason": "correo_no_aprobado"})
                continue

            if not OUT_BUCKET:
                skipped.append({"key": in_key, "correo": correo, "reason": "OUT_BUCKET_no_configurado"})
                continue

            s3_uri = f"s3://{in_bucket}/{in_key}"
            job_name = sanitize_job_name(in_key)
            media_format = in_key.split(".")[-1].lower()

            output_path = _build_output_path()

            params: Dict[str, Any] = {
                "TranscriptionJobName": job_name,
                "Media": {"MediaFileUri": s3_uri},
                "MediaFormat": media_format,
                "OutputBucketName": OUT_BUCKET,
                "OutputKey": output_path + "/",
            }

            if LANGUAGE == "auto":
                params["IdentifyLanguage"] = True
            else:
                params["LanguageCode"] = LANGUAGE

            if SPEAKERS >= 2:
                params["Settings"] = {"ShowSpeakerLabels": True, "MaxSpeakerLabels": SPEAKERS}

            try:
                transcribe.start_transcription_job(**params)
                processed.append({"key": in_key, "correo": correo, "job": job_name, "output": f"s3://{OUT_BUCKET}/{output_path}/"})
            except ClientError as e:
                code = e.response.get("Error", {}).get("Code")
                if code == "ConflictException":
                    processed.append({"key": in_key, "correo": correo, "job": job_name, "status": "already_exists"})
                else:
                    raise

        except Exception as e:
            skipped.append({"reason": "exception", "error": str(e), "record": record})

    return {
        "statusCode": 200,
        "body": json.dumps({"processed": processed, "skipped": skipped}, ensure_ascii=False),
    }
